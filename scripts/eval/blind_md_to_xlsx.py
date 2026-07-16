#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
blind_md_to_xlsx —— 把重排「盲标表.md」+「gold-scaffold.json」合成人工标注用 xlsx。
  一行 = 一个 (故事 × 候选题) 对；金标档下拉校验；故事全文每行都带；冻结表头；填写列上色。
  严禁出现 AI 分数 / AI 理由 / zone / 召回来源 —— 盲标。questionId 置末列，仅供回写合并。
用法：
  python3 scripts/eval/blind_md_to_xlsx.py \
      --md    scripts/eval/results/ranking-XXXX-盲标表.md \
      --scaffold scripts/eval/results/ranking-XXXX-gold-scaffold.json \
      --out   scripts/eval/results/ranking-XXXX-盲标表.xlsx
依赖：openpyxl（系统 python，不进项目依赖）。
"""
import argparse, json, re, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

TIERS = ['高', '中', '低', '隐藏']

# ── 解析盲标 md ───────────────────────────────────────────────────────────────

def parse_md(md_text):
    """返回 [{storyId, story, rows:[{part,en,zh,point}]}]，rows 保持 md 中的中性顺序。"""
    stories = []
    cur = None
    lines = md_text.splitlines()
    for line in lines:
        m = re.match(r'^##\s+(S\d+)', line)
        if m:
            if cur is not None:
                stories.append(cur)
            cur = {'storyId': m.group(1), 'story': '', 'rows': []}
            continue
        if cur is None:
            continue
        if line.startswith('- 故事全文：'):
            cur['story'] = line.split('：', 1)[1].strip()
            continue
        # 表数据行：以 | 开头、非表头(| Part)、非分隔(|---)
        if line.startswith('|') and not line.startswith('| Part') and not re.match(r'^\|[-\s|]+\|?$', line):
            cells = [c.strip().replace('\\|', '|') for c in line.split('|')]
            # 形如 ['', part, en, zh, point, 金标档(空), 备注(空), '']
            if len(cells) >= 7:
                cur['rows'].append({'part': cells[1], 'en': cells[2], 'zh': cells[3], 'point': cells[4]})
    if cur is not None:
        stories.append(cur)
    return stories

# ── 说明 / 速查 文案 ──────────────────────────────────────────────────────────

GUIDE_TITLE = '重排金标 · 盲标说明'
GUIDE_BLOCKS = [
    ('怎么标', [
        '一行 = 一个「故事 × 候选题」对。你只做一件事：在「金标档」列的下拉里选 高/中/低/隐藏。',
        '⚠️ 绿色「金标档」= 从上版 ranking.v1 继承（题面未变），已填好，别动；只需标黄色空格那些行。',
        '判据只有一条：拿「故事全文」这段经历，能不能自然、充分地回答这道候选题——跟这题怎么来的、AI 怎么想的无关（本表刻意不给这些）。',
        '拿不准就在「备注」写一句你的纠结点。逐题独立判，别因为同一个故事的另一道题怎么判就带着判这道。',
        '最后一列 questionId 是回写用的机器编号，别动、别删、别排序打乱行序。',
    ]),
    ('四档定义（改不改故事）', [
        '高（原样能答）：故事的重心 / 主语 / 场景 / 时间 / 活动 / 聚光灯落点【全不用改】，照现有故事就能直接、充分回答，而且答的正是这道题在问的点。',
        '中（改角度能答）：沾边，但要换角度 / 挪聚光灯 / 换场景 / 只覆盖题目一部分 / 把「日常习惯」硬套成「某一次」才能答；故事主体还能用，无需另起一个完全不同的故事。',
        '低（得换故事，但同域·可露出）：用现有故事答不了，必须换一个不同的经历；但题目与故事【同域 / 沾边】，摆进「查看更多」用户看得懂关联。',
        '隐藏（跨域无关·不露出）：题目与故事【跨域、无交集】，用现有故事答会答非所问；露出来用户会困惑「这题跟我讲的有什么关系」。',
    ]),
    ('★ 低 vs 隐藏 —— 最要紧的界（它就是"展示 vs 不展示"的切分）', [
        '共同点：低和隐藏，用现有故事都【答不了】（都得换个经历）。区别不在"能不能答"，而在"露出来合不合理"。',
        '低：题目与故事【同域/沾边】，用户第一反应"嗯还算相关，只是我这次没这段经历" → 值得作为"查看更多"折叠露出。',
        '隐藏：题目与故事【跨域、无交集】，用户第一反应"这题怎么跑这儿来了？" → 露出即噪声，不展示。',
        '决胜问句：把这道题摆进"查看更多"列表，用户会觉得"还算相关"（→低）还是"这题怎么在这"（→隐藏）？',
        '锚点例：咖啡放松的故事 × "你更喜欢打字还是手写？" = 跨域无交集 → 隐藏。',
        '纠结点（同类事件、内容不搭）：如"赶论文崩溃"故事 × "迷路的一次经历" —— 同属"某一次具体事件"（同域），但内容不搭、得换故事 → 判【低】（同域沾边、露出不突兀），不判隐藏。',
    ]),
]

LOOKUP_ROWS = [
    ('高', '原样能答', '重心/主语/场景/时间/活动/聚光灯全不用改，直接充分答，且正答它问的点'),
    ('中', '改角度能答', '沾边，但要换角度/挪聚光灯/换场景/只覆盖一部分/习惯硬套成某一次'),
    ('低', '得换故事·同域可露', '用现有故事答不了，须换经历；但与故事同域/沾边，露进"查看更多"看得懂关联'),
    ('隐藏', '跨域无关·不露', '与故事跨域、无交集，答会答非所问；露出即噪声'),
]
LOOKUP_RULES = [
    '高/中界：聚光灯 / 重心要不要挪？要挪即降中。原样答的正是它问的，才给高。',
    '中/低界：换个角度或侧重、还能用这故事答 = 中；必须换一个完全不同的经历 = 低。',
    '低/隐藏界：同域/沾边、露出用户看得懂关联 = 低；跨域无交集、露出是噪声 = 隐藏。',
    '习惯 vs 某一次：故事讲常态、题问"某一次"，要硬套 = 中。',
    '场景没发生：题目要求的场景/活动故事里压根没出现 → 至少低；若还跨域则隐藏。',
    '跨语言：故事中文、题目英文，按语义判，别因语言不同误判。',
    '逐题独立：别因为同故事另一道题怎么判，就带着判这道。',
]

# ── 样式 ──────────────────────────────────────────────────────────────────────

HEAD_FILL = PatternFill('solid', fgColor='305496')
HEAD_FONT = Font(color='FFFFFF', bold=True)
FILLME_FILL = PatternFill('solid', fgColor='FFF2CC')   # 待填列：浅黄
QID_FILL = PatternFill('solid', fgColor='EDEDED')       # questionId：浅灰=别动
INHERIT_FILL = PatternFill('solid', fgColor='C6EFCE')   # 继承已填：浅绿=别动
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
THIN = Side(style='thin', color='D9D9D9')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def build_workbook(stories, scaffold):
    # storyId -> ordered labels（questionId 顺序与 md 行序一致，均出自同一 selectForLabeling）
    scaf_by_story = {it['storyId']: it['labels'] for it in scaffold['items']}

    wb = Workbook()

    # ── Sheet 1：标注说明 ──
    ws1 = wb.active
    ws1.title = '标注说明'
    ws1.column_dimensions['A'].width = 4
    ws1.column_dimensions['B'].width = 120
    r = 1
    ws1.cell(r, 2, GUIDE_TITLE).font = Font(bold=True, size=14); r += 2
    for head, items in GUIDE_BLOCKS:
        ws1.cell(r, 2, head).font = Font(bold=True, size=12, color='C00000' if head.startswith('★') else '305496'); r += 1
        for it in items:
            c = ws1.cell(r, 2, '· ' + it); c.alignment = WRAP; r += 1
        r += 1

    # ── Sheet 2：标注表 ──
    ws = wb.create_sheet('标注表')
    headers = ['故事编号', '故事全文', 'Part', '题目(英文)', '题目(中文)', '所属观察点', '金标档', '备注', 'questionId']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h); c.fill = HEAD_FILL; c.font = HEAD_FONT; c.alignment = CENTER; c.border = BORDER
    widths = [10, 52, 6, 42, 30, 20, 10, 26, 40]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + ci)].width = w
    ws.freeze_panes = 'A2'

    row = 2
    skipped = []
    for st in stories:
        labels = scaf_by_story.get(st['storyId'], [])
        if len(labels) != len(st['rows']):
            skipped.append(f"{st['storyId']}(md {len(st['rows'])} 行 vs scaffold {len(labels)} 标)")
            continue
        for md_row, lab in zip(st['rows'], labels):
            # 校验行序对齐：part 应一致
            if str(lab.get('part')) != str(md_row['part']):
                skipped.append(f"{st['storyId']} 行序错位(part {md_row['part']}≠{lab.get('part')})")
                break
            inherited = lab.get('goldBucket')
            note = '继承自 ranking.v1（题面未变，勿改）' if inherited else ''
            vals = [st['storyId'], st['story'], md_row['part'], md_row['en'], md_row['zh'], md_row['point'], inherited or '', note, lab['questionId']]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row, ci, v); c.border = BORDER
                if ci in (2, 4, 5, 6):
                    c.alignment = WRAP
                elif ci == 3:
                    c.alignment = CENTER
                elif ci == 7:
                    c.fill = INHERIT_FILL if inherited else FILLME_FILL; c.alignment = CENTER
                elif ci == 8:
                    c.fill = INHERIT_FILL if inherited else FILLME_FILL; c.alignment = WRAP
                    if inherited:
                        c.font = Font(size=9, color='808080')
                elif ci == 9:
                    c.fill = QID_FILL; c.font = Font(size=9, color='808080'); c.alignment = Alignment(vertical='top')
                else:
                    c.alignment = Alignment(vertical='top')
            row += 1
    last = row - 1

    # 金标档下拉校验（G 列）
    dv = DataValidation(type='list', formula1='"高,中,低,隐藏"', allow_blank=True)
    dv.showErrorMessage = True
    dv.errorTitle = '档位无效'
    dv.error = '只能从下拉选择：高 / 中 / 低 / 隐藏'
    dv.promptTitle = '选金标档'
    dv.prompt = '拿这故事能不能答这道题：高/中/低/隐藏'
    dv.showInputMessage = True
    ws.add_data_validation(dv)
    if last >= 2:
        dv.add(f'G2:G{last}')

    # ── Sheet 3：档位速查 ──
    ws3 = wb.create_sheet('档位速查')
    ws3.column_dimensions['A'].width = 8
    ws3.column_dimensions['B'].width = 20
    ws3.column_dimensions['C'].width = 90
    ws3.cell(1, 1, '四档定义').font = Font(bold=True, size=13)
    for ci, h in enumerate(['档', '一句话', '判据'], 1):
        c = ws3.cell(2, ci, h); c.fill = HEAD_FILL; c.font = HEAD_FONT; c.alignment = CENTER
    rr = 3
    for tier, short, desc in LOOKUP_ROWS:
        ws3.cell(rr, 1, tier).alignment = CENTER
        ws3.cell(rr, 2, short).alignment = WRAP
        ws3.cell(rr, 3, desc).alignment = WRAP
        rr += 1
    rr += 1
    ws3.cell(rr, 1, '边界仲裁细则').font = Font(bold=True, size=13); rr += 1
    for rule in LOOKUP_RULES:
        c = ws3.cell(rr, 1, '· ' + rule); ws3.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=3)
        c.alignment = WRAP; rr += 1

    return wb, last, skipped


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--md', required=True)
    ap.add_argument('--scaffold', required=True)
    ap.add_argument('--out', required=True)
    args = ap.parse_args()

    with open(args.md, encoding='utf-8') as f:
        stories = parse_md(f.read())
    with open(args.scaffold, encoding='utf-8') as f:
        scaffold = json.load(f)

    wb, last, skipped = build_workbook(stories, scaffold)
    wb.save(args.out)

    print(f'✓ 生成 {args.out}')
    print(f'  故事 {len(stories)} 条｜标注行 {last - 1} 行')
    if skipped:
        print('  ⚠ 跳过/告警：', '; '.join(skipped))
    else:
        print('  行序与 scaffold 完全对齐（part 校验通过）')


if __name__ == '__main__':
    sys.exit(main())

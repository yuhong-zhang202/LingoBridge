#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
xlsx_to_gold —— 读「填完的盲标 xlsx」→ 按 questionId 合并进 gold-scaffold.json → 产出金标 ranking.v1.json。
  scaffold 提供骨架与 zone；xlsx 提供人标的 goldBucket（+备注）。二者以 questionId 为准对齐。
  ⚠ 这是「造考卷」：金标是评判 AI 的标准答案，须主会话亲自跑并核对，不可交给下游改代码的 agent。
用法：
  python3 scripts/eval/xlsx_to_gold.py \
      --xlsx     scripts/eval/results/ranking-XXXX-盲标表.xlsx \
      --scaffold scripts/eval/results/ranking-XXXX-gold-scaffold.json \
      --out      scripts/eval/golden/ranking.v1.json \
      --annotator "张三&李四" [--annotated-at 2026-07-16] [--allow-incomplete]
不完整（有未填/非法档位）默认拒绝出金标，除非显式 --allow-incomplete。
"""
import argparse, json, os, sys, datetime
from openpyxl import load_workbook

TIERS = {'高', '中', '低', '隐藏'}


def read_xlsx(path):
    """返回 {(storyId, questionId): {'bucket': str|None, 'note': str}}，及重复键列表。
    注意：同一 questionId 可被多个故事召回、且各故事应判不同档，故键必须含 storyId。"""
    wb = load_workbook(path, data_only=True)
    if '标注表' not in wb.sheetnames:
        sys.exit(f'✗ xlsx 缺少「标注表」sheet：{path}')
    ws = wb['标注表']
    hdr = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    try:
        si = hdr.index('故事编号') + 1
        gi = hdr.index('金标档') + 1
        ni = hdr.index('备注') + 1
        qi = hdr.index('questionId') + 1
    except ValueError as e:
        sys.exit(f'✗ 表头缺列：{e}（需要 故事编号 / 金标档 / 备注 / questionId）')

    out = {}
    dups = []
    for r in range(2, ws.max_row + 1):
        qid = ws.cell(r, qi).value
        sid = ws.cell(r, si).value
        if qid in (None, '') or sid in (None, ''):
            continue
        key = (str(sid).strip(), str(qid).strip())
        bucket = ws.cell(r, gi).value
        bucket = str(bucket).strip() if bucket not in (None, '') else None
        note = ws.cell(r, ni).value
        note = str(note).strip() if note not in (None, '') else ''
        if key in out:
            dups.append(key)
        out[key] = {'bucket': bucket, 'note': note}
    return out, dups


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--xlsx', required=True)
    ap.add_argument('--scaffold', required=True)
    ap.add_argument('--out', default='scripts/eval/golden/ranking.v1.json')
    ap.add_argument('--annotator', required=True)
    ap.add_argument('--annotated-at', default=datetime.date.today().isoformat())
    ap.add_argument('--allow-incomplete', action='store_true')
    args = ap.parse_args()

    labels_map, dups = read_xlsx(args.xlsx)
    with open(args.scaffold, encoding='utf-8') as f:
        scaffold = json.load(f)

    # 对齐核对（键 = (storyId, questionId)）
    scaf_keys = {(it['storyId'], lab['questionId']) for it in scaffold['items'] for lab in it['labels']}
    xlsx_keys = set(labels_map)
    missing_in_xlsx = scaf_keys - xlsx_keys          # 骨架有、表里没 → 行被删了？
    extra_in_xlsx = xlsx_keys - scaf_keys            # 表里有、骨架没 → 串了别的文件
    both = scaf_keys & xlsx_keys
    unfilled = [k for k in both if labels_map[k]['bucket'] is None]
    invalid = [(k, labels_map[k]['bucket']) for k in both
               if labels_map[k]['bucket'] is not None and labels_map[k]['bucket'] not in TIERS]

    problems = []
    if dups:
        problems.append(f'questionId 重复 {len(dups)} 个：{dups[:5]}...')
    if missing_in_xlsx:
        problems.append(f'骨架有、xlsx 缺 {len(missing_in_xlsx)} 行（行被删？）：{list(missing_in_xlsx)[:5]}...')
    if extra_in_xlsx:
        problems.append(f'xlsx 多出 {len(extra_in_xlsx)} 个 questionId（文件对错了？）：{list(extra_in_xlsx)[:5]}...')
    if invalid:
        problems.append(f'非法档位 {len(invalid)} 处：{invalid[:5]}...')
    if unfilled:
        problems.append(f'未填档位 {len(unfilled)} 处：{unfilled[:8]}...')

    if problems and not args.allow_incomplete:
        print('✗ 金标未产出——先修下列问题（或加 --allow-incomplete 强出，不推荐）：')
        for p in problems:
            print('  -', p)
        sys.exit(1)
    if problems:
        print('⚠ --allow-incomplete：带下列问题强行产出，未填项 goldBucket=null：')
        for p in problems:
            print('  -', p)

    # 合并：把 bucket/note 填进骨架
    tier_dist = {t: 0 for t in TIERS}
    zone_dist = {'visible': 0, 'hidden_sampled': 0}
    for it in scaffold['items']:
        for lab in it['labels']:
            m = labels_map.get((it['storyId'], lab['questionId']))
            lab['goldBucket'] = m['bucket'] if m else None
            if m and m['note']:
                lab['note'] = m['note']
            elif 'note' in lab:
                del lab['note']
            if lab['goldBucket'] in tier_dist:
                tier_dist[lab['goldBucket']] += 1
            if lab.get('zone') in zone_dist:
                zone_dist[lab['zone']] += 1

    gold = {
        'version': scaffold.get('version', 'ranking-v1'),
        'bucketRubric': scaffold.get('bucketRubric', ''),
        'hiddenSampleRate': scaffold.get('hiddenSampleRate'),
        'annotator': args.annotator,
        'annotatedAt': args.annotated_at,
        'items': scaffold['items'],
    }

    os.makedirs(os.path.dirname(args.out), exist_ok=True)
    with open(args.out, 'w', encoding='utf-8') as f:
        json.dump(gold, f, ensure_ascii=False, indent=2)

    total = sum(len(it['labels']) for it in scaffold['items'])
    print(f'✓ 金标产出 {args.out}')
    print(f'  故事 {len(scaffold["items"])} 条｜标注对 {total} 个')
    print(f'  档位分布：' + '｜'.join(f'{t} {tier_dist[t]}' for t in ['高', '中', '低', '隐藏']))
    print(f'  zone：visible {zone_dist["visible"]}｜hidden_sampled {zone_dist["hidden_sampled"]}')
    print(f'  标注人 {args.annotator}｜日期 {args.annotated_at}')
    print('  下一步：npm run eval:ranking:score -- --export=scripts/eval/results/<对应导出>.json')


if __name__ == '__main__':
    sys.exit(main())
